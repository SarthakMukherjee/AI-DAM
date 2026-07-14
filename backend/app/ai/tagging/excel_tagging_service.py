import openpyxl

class ExcelTaggingService:
    def extract_text(self, excel_path: str) -> str:
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
            full_text = []
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    row_text = []
                    for cell in row:
                        if cell is not None:
                            row_text.append(str(cell).strip())
                    if row_text:
                        full_text.append(" | ".join(row_text))
            wb.close()
            return "\n".join(full_text).strip()
        except Exception as e:
            print(f"Excel text extraction failed: {e}")
            return ""

    def extract_metadata(self, excel_path: str) -> dict:
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
            props = wb.properties
            meta = {
                "title": props.title or "",
                "author": props.creator or "",
                "subject": props.subject or "",
                "keywords": props.keywords or ""
            }
            wb.close()
            return meta
        except Exception as e:
            print(f"Excel metadata extraction failed: {e}")
            return {
                "title": "",
                "author": "",
                "subject": "",
                "keywords": ""
            }
